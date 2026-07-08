import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('招聘与入职流程表.xlsx', data_only=True)
    print('Sheets:', wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'Dimensions: {ws.dimensions}')
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
        
        # 打印前30行
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append((cell.column_letter, cell.value))
            if row_data:
                print(f'Row {row_idx}: {row_data}')
                
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()
